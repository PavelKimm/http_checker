import socket
import requests
import openpyxl

from django.conf import settings
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction, OperationalError
from django.db.models import Q
from rest_framework import generics
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.generics import get_object_or_404
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.status import (
    HTTP_200_OK, HTTP_204_NO_CONTENT, HTTP_400_BAD_REQUEST, HTTP_401_UNAUTHORIZED,
    HTTP_403_FORBIDDEN, HTTP_404_NOT_FOUND, HTTP_500_INTERNAL_SERVER_ERROR, HTTP_201_CREATED
)

from analyzer.models import Website, WebsiteAvailability
from analyzer.serializers import WebsiteAvailabilityDetailSerializer

START_ROW = 2  # starting row in excel file
TIMEOUT = 10  # waiting for the website responds


@api_view(['POST'])
def load_websites_to_check(request):
    """
    insertion of website urls to the database from .xlsx files
    """

    inserted_urls_count = 0
    passed_urls_count = 0

    xlsx_files = request.FILES.getlist('files')
    for xlsx_file in xlsx_files:
        try:
            wb_obj = openpyxl.load_workbook(xlsx_file)
        except Exception as e:
            return Response({"detail": str(e)}, status=HTTP_400_BAD_REQUEST)

        # if you need to get specific sheet
        # first_sheet = wb_obj.get_sheet_names()[0]
        # sheet = wb_obj.get_sheet_by_name(first_sheet)

        sheet = wb_obj.active

        for row in sheet.iter_rows(min_row=START_ROW, max_row=sheet.max_row):
            for cell in row:
                try:
                    Website.objects.create(url=cell.value.strip())
                    inserted_urls_count += 1
                except Exception as e:
                    # print(str(e))
                    passed_urls_count += 1

    return Response(f"The insertion has been completed. Inserted urls number: {inserted_urls_count}. "
                    f"Passed urls number: {passed_urls_count}", status=HTTP_200_OK)


@api_view(['POST'])
def check_websites(request):
    """
    check websites availability with logging to the database
    """

    start_from_id = request.data.get('start_from_id')
    checked_websites_count = 0
    failed_websites = []

    websites = Website.objects.order_by('id')
    for website in websites:
        if start_from_id:
            if start_from_id > website.id:
                continue
        print(website.url)
        # raw_url == urn
        raw_url = website.url
        if raw_url.startswith("http://") or raw_url.startswith("https://"):
            url = raw_url
            raw_url = raw_url.split("://")[-1]
        else:
            url = "http://" + raw_url

        try:
            res = requests.get(url, timeout=TIMEOUT)

            response_status_code = res.status_code
            reason = res.reason
            server = res.headers.get('Server')
            ip_address = socket.gethostbyname(raw_url)
            print(ip_address)
            WebsiteAvailability.objects.create(website=website, response_status_code=response_status_code,
                                               reason=reason, server=server, ip_address=ip_address)

            checked_websites_count += 1
        except OperationalError as e:
            return Response({"detail": str(e)}, status=HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            WebsiteAvailability.objects.create(website=website, reason=str(e))
            failed_websites.append({
                "id": website.id,
                "url": website.url
            })
    print(failed_websites)
    return Response(f"Checked websites number: {checked_websites_count}", status=HTTP_200_OK)


@api_view(['GET'])
def get_website_info_by_url(request):
    """
    get website check information by url
    """
    url = request.query_params.get('url')
    try:
        # getting information of the last check
        check_info = WebsiteAvailability.objects.filter(website__url=url).order_by('-checked_at').first()
    except WebsiteAvailability.DoesNotExist as e:
        return Response({"detail": str(e)}, status=HTTP_404_NOT_FOUND)
    serializer = WebsiteAvailabilityDetailSerializer(check_info)
    return Response(serializer.data, status=HTTP_200_OK)
